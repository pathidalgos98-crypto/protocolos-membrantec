from flask import Flask, request, send_file, jsonify
from flask_cors import CORS
import openpyxl
from io import BytesIO
import os
from datetime import datetime, time as dtime

app = Flask(__name__)
CORS(app)

PLANTILLAS = os.path.join(os.path.dirname(__file__), 'Aplicación Membrantec')
print(f'[membrantec] Plantillas en: {PLANTILLAS}')

ARCHIVOS = {
    'pi':  'GM-PI-HDPE-INF-RAI-001.xlsx',
    'pd':  'GM-PD-HDPE-INF-RAI-001.xlsx',
    'ru':  'GM-RU-HDPE-INF-RAI-001.xlsx',
    'rep': 'GM-REP-HDPE-SUP-RAI-001_1.xlsx',
    'ri':  'GM-RI-HDPE-SUP-RAI-001.xlsx',
}

def to_date(val):
    if not val: return None
    try: return datetime.strptime(str(val)[:10], '%Y-%m-%d')
    except: return val

def to_time(val):
    if not val: return None
    try:
        partes = str(val).replace('.', ':').split(':')
        return dtime(int(partes[0]), int(partes[1]))
    except: return val

@app.route('/')
def index():
    return jsonify({'status': 'GEOMEMB API OK'})

@app.route('/generar-protocolo', methods=['POST', 'OPTIONS'])
def generar_protocolo():
    if request.method == 'OPTIONS':
        return '', 204

    try:
        d = request.json
        tipo = d.get('tipo', '').lower()
        proy = d.get('proyecto', {})
        registros = d.get('registros', [])
        manometro = d.get('manometro', {})
        equipo = d.get('equipo', {})

        if tipo not in ARCHIVOS:
            return jsonify({'error': f'Tipo desconocido: {tipo}'}), 400

        ruta = os.path.join(PLANTILLAS, ARCHIVOS[tipo])
        if not os.path.exists(ruta):
            return jsonify({'error': f'Plantilla no encontrada: {ruta}'}), 500

        wb = openpyxl.load_workbook(ruta)
        ws = wb.active
        proto = proy.get('protocolo', tipo.upper())

        if tipo == 'pi':
            ws['F4']  = f"Contrato : {proy.get('contrato','')}"
            ws['F5']  = f"Proyecto: \"{proy.get('proyecto','')}\""
            ws['D7']  = proy.get('area_cod','')
            ws['J7']  = proy.get('plano','')
            ws['S7']  = proto
            ws['AB7'] = to_date(proy.get('fecha',''))
            ws['A9']  = proy.get('tag','')
            ws['C10'] = proy.get('area_cod','')
            ws['E10'] = proy.get('area_desc','')
            ws['H10'] = proy.get('sis_cod','')
            ws['K10'] = proy.get('sis_desc','')
            ws['M10'] = proy.get('sub_cod','')
            ws['P10'] = proy.get('sub_desc','')
            ws['S10'] = proy.get('aconex','')
            ws['Y10'] = proy.get('pie','')
            ws['C25'] = proy.get('min_des', 121)
            ws['E25'] = proy.get('min_cor', 160)
            ws['C28'] = proy.get('tens_eq', 'TS-25')
            ws['C29'] = proy.get('tens_cert', 'MAM-3494')
            fila = 36
            for r in registros:
                ws[f'A{fila}'] = to_date(r.get('fecha',''))
                ws[f'C{fila}'] = to_time(r.get('hora',''))
                ws[f'D{fila}'] = r.get('operador','')
                ws[f'E{fila}'] = r.get('tamb','')
                ws[f'F{fila}'] = r.get('maq','')
                ws[f'H{fila}'] = r.get('tmaq','')
                ws[f'J{fila}'] = r.get('tleister','-')
                ws[f'K{fila}'] = r.get('veloc','')
                ws[f'L{fila}'] = r.get('inspector','')
                ws[f'N{fila}'] = r.get('d1a',''); ws[f'O{fila}'] = r.get('d1b','')
                ws[f'P{fila}'] = r.get('d2a',''); ws[f'Q{fila}'] = r.get('d2b','')
                ws[f'R{fila}'] = r.get('d3a',''); ws[f'S{fila}'] = r.get('d3b','')
                ws[f'T{fila}'] = r.get('d4a',''); ws[f'U{fila}'] = r.get('d4b','')
                ws[f'V{fila}'] = r.get('d5a',''); ws[f'W{fila}'] = r.get('d5b','')
                ws[f'X{fila}'] = r.get('c1',''); ws[f'Y{fila}'] = r.get('c2','')
                ws[f'Z{fila}'] = r.get('c3',''); ws[f'AA{fila}'] = r.get('c4','')
                ws[f'AB{fila}'] = r.get('c5','')
                ws[f'AC{fila}'] = r.get('resultado','')
                ws[f'AD{fila}'] = r.get('falla_prob','-')
                ws[f'AE{fila}'] = r.get('falla_tipo','-')
                fila += 1

        elif tipo == 'pd':
            ws['E4']  = f"Contrato : {proy.get('contrato','')}"
            ws['E5']  = f"Proyecto: \"{proy.get('proyecto','')}\""
            ws['D7']  = proy.get('area_cod','')
            ws['J7']  = proy.get('plano','')
            ws['S7']  = proto
            ws['AB7'] = to_date(proy.get('fecha',''))
            ws['A10'] = proy.get('tag','')
            ws['D10'] = proy.get('area_cod','')
            ws['F10'] = proy.get('area_desc','')
            ws['I10'] = proy.get('sis_cod','')
            ws['L10'] = proy.get('sis_desc','')
            ws['P10'] = proy.get('sub_cod','')
            ws['S10'] = proy.get('sub_desc','')
            ws['V10'] = proy.get('aconex','')
            ws['AB10'] = proy.get('pie','')
            ws['C25'] = proy.get('min_des', 121)
            ws['E25'] = proy.get('min_cor', 160)
            ws['D27'] = proy.get('tens_eq', 'TS-25')
            ws['J27'] = proy.get('tens_cert', 'MAM-3494')
            fila = 35
            for i, r in enumerate(registros):
                ws[f'A{fila}'] = to_date(r.get('fecha',''))
                ws[f'B{fila}'] = i + 1
                ws[f'C{fila}'] = r.get('union','')
                ws[f'D{fila}'] = r.get('ubicacion','')
                ws[f'E{fila}'] = r.get('hora','')
                ws[f'F{fila}'] = r.get('tamb','')
                ws[f'G{fila}'] = r.get('operador','')
                ws[f'H{fila}'] = r.get('maq','')
                ws[f'I{fila}'] = r.get('tmaq','')
                ws[f'J{fila}'] = r.get('tleister','-')
                ws[f'K{fila}'] = r.get('veloc','')
                ws[f'L{fila}'] = r.get('inspector','')
                ws[f'N{fila}'] = r.get('d1a',''); ws[f'O{fila}'] = r.get('d1b','')
                ws[f'P{fila}'] = r.get('d2a',''); ws[f'Q{fila}'] = r.get('d2b','')
                ws[f'R{fila}'] = r.get('d3a',''); ws[f'S{fila}'] = r.get('d3b','')
                ws[f'T{fila}'] = r.get('d4a',''); ws[f'U{fila}'] = r.get('d4b','')
                ws[f'V{fila}'] = r.get('d5a',''); ws[f'W{fila}'] = r.get('d5b','')
                ws[f'X{fila}'] = r.get('c1',''); ws[f'Y{fila}'] = r.get('c2','')
                ws[f'Z{fila}'] = r.get('c3',''); ws[f'AA{fila}'] = r.get('c4','')
                ws[f'AB{fila}'] = r.get('c5','')
                ws[f'AC{fila}'] = r.get('resultado','')
                ws[f'AD{fila}'] = r.get('falla_prob','-')
                ws[f'AE{fila}'] = r.get('falla_tipo','-')
                fila += 1

        elif tipo == 'ru':
            ws['E3']  = f"Contrato : {proy.get('contrato','')}"
            ws['E4']  = f"Proyecto: \"{proy.get('proyecto','')}\""
            ws['D6']  = proy.get('area_cod','')
            ws['H6']  = proy.get('plano','')
            ws['P6']  = proto
            ws['V6']  = to_date(proy.get('fecha',''))
            ws['B8']  = proy.get('tag','')
            ws['D9']  = proy.get('area_cod','')
            ws['E9']  = proy.get('area_desc','')
            ws['H9']  = proy.get('sis_cod','')
            ws['K9']  = proy.get('sis_desc','')
            ws['O9']  = proy.get('sub_cod','')
            ws['Q9']  = proy.get('sub_desc','')
            ws['S9']  = proy.get('aconex','')
            ws['U9']  = proy.get('pie','')
            ws['E12'] = proy.get('lugar','')
            ws['N39'] = manometro.get('serie','')
            ws['Q39'] = to_date(manometro.get('fecha_cal',''))
            ws['T39'] = manometro.get('certificado','')
            ws['V39'] = manometro.get('etiqueta','')
            fila = 26
            for r in registros:
                ws[f'B{fila}'] = r.get('union','')
                ws[f'C{fila}'] = r.get('distancia','')
                ws[f'D{fila}'] = to_date(r.get('fecha',''))
                ws[f'E{fila}'] = r.get('operador','')
                ws[f'F{fila}'] = r.get('maq','')
                ws[f'G{fila}'] = r.get('tmaq','')
                ws[f'H{fila}'] = r.get('veloc','')
                ws[f'I{fila}'] = to_time(r.get('hora_union',''))
                ws[f'J{fila}'] = 'X'
                ws[f'K{fila}'] = '-'
                ws[f'L{fila}'] = r.get('ds', 0)
                ws[f'N{fila}'] = to_time(r.get('hora_ini',''))
                ws[f'O{fila}'] = r.get('psi_ini','')
                ws[f'P{fila}'] = to_time(r.get('hora_fin',''))
                ws[f'Q{fila}'] = r.get('psi_fin','')
                ws[f'R{fila}'] = r.get('psi_dif','')
                ws[f'S{fila}'] = r.get('resultado','')
                ws[f'T{fila}'] = r.get('tec','')
                ws[f'U{fila}'] = to_date(r.get('fecha_prueba',''))
                ws[f'V{fila}'] = r.get('tipo_falla','-')
                ws[f'W{fila}'] = r.get('mano_etiq','')
                fila += 1

        elif tipo == 'rep':
            ws['G3']  = f"Contrato : {proy.get('contrato','')}"
            ws['G4']  = f"Proyecto: \"{proy.get('proyecto','')}\""
            ws['D6']  = proy.get('area_cod','')
            ws['I6']  = proy.get('plano','')
            ws['Q6']  = proto
            ws['X6']  = to_date(proy.get('fecha',''))
            ws['B9']  = proy.get('tag','')
            ws['D9']  = proy.get('area_cod','')
            ws['E9']  = proy.get('area_desc','')
            ws['G9']  = proy.get('sis_cod','')
            ws['H9']  = proy.get('sis_desc','')
            ws['J9']  = proy.get('sub_cod','')
            ws['L9']  = proy.get('sub_desc','')
            ws['N9']  = proy.get('aconex','')
            ws['U9']  = proy.get('pie','')
            ws['E11'] = proy.get('lugar','')
            ws['G18'] = equipo.get('nombre','VAC-26')
            ws['J18'] = to_date(equipo.get('fecha_cal',''))
            ws['L18'] = equipo.get('certificado','')
            fila = 27
            for r in registros:
                ws[f'B{fila}'] = to_date(r.get('fecha',''))
                ws[f'C{fila}'] = to_time(r.get('hora',''))
                ws[f'D{fila}'] = r.get('n_parche','')
                ws[f'E{fila}'] = r.get('tipo_rep','')
                ws[f'F{fila}'] = r.get('union','')
                ws[f'H{fila}'] = r.get('tecnico','')
                ws[f'J{fila}'] = r.get('dim_largo','')
                ws[f'K{fila}'] = r.get('dim_ancho','')
                ws[f'L{fila}'] = r.get('maq','')
                ws[f'M{fila}'] = r.get('tmaq','')
                ws[f'N{fila}'] = r.get('tleister','')
                ws[f'O{fila}'] = r.get('ds','-')
                ws[f'Q{fila}'] = r.get('prueba','VA')
                ws[f'S{fila}'] = '✓' if r.get('resultado') == 'Aprobado' else '✗'
                ws[f'U{fila}'] = r.get('tipo_falla1','-')
                ws[f'V{fila}'] = r.get('equipo','')
                ws[f'W{fila}'] = to_time(r.get('hora_fin',''))
                ws[f'X{fila}'] = to_date(r.get('fecha_fin',''))
                ws[f'Y{fila}'] = r.get('op_fin','')
                fila += 1

        elif tipo == 'ri':
            ws['G4']  = f"Contrato : {proy.get('contrato','')}"
            ws['G5']  = f"Proyecto: \"{proy.get('proyecto','')}\""
            ws['E7']  = proy.get('area_cod','')
            ws['M7']  = proy.get('plano','')
            ws['U7']  = proto
            ws['AB7'] = to_date(proy.get('fecha',''))
            ws['B9']  = proy.get('tag','')
            ws['E10'] = proy.get('area_cod','')
            ws['I10'] = proy.get('area_desc','')
            ws['L10'] = proy.get('sis_cod','')
            ws['O10'] = proy.get('sis_desc','')
            ws['R10'] = proy.get('sub_cod','')
            ws['T10'] = proy.get('sub_desc','')
            ws['V10'] = proy.get('aconex','')
            ws['AB10'] = proy.get('pie','')
            ws['F13'] = proy.get('lugar','')
            fila = 22
            for r in registros:
                clima = r.get('clima','')
                ws[f'B{fila}'] = to_date(r.get('fecha',''))
                ws[f'D{fila}'] = to_time(r.get('hora',''))
                ws[f'E{fila}'] = 'x' if clima == 'Nublado' else '-'
                ws[f'F{fila}'] = 'x' if clima == 'Despejado c/viento' else '-'
                ws[f'G{fila}'] = 'x' if clima == 'Despejado s/viento' else '-'
                ws[f'I{fila}'] = r.get('panel','')
                ws[f'K{fila}'] = r.get('rollo','')
                ws[f'M{fila}'] = r.get('m2_rollo','')
                ws[f'P{fila}'] = r.get('espesor','2.00')
                ws[f'Q{fila}'] = r.get('st_largo','')
                ws[f'R{fila}'] = r.get('st_ancho','')
                ws[f'S{fila}'] = r.get('st_m2','')
                ws[f'U{fila}'] = r.get('ct_largo','')
                ws[f'V{fila}'] = r.get('ct_ancho','')
                ws[f'W{fila}'] = r.get('ct_m2','')
                ws[f'Z{fila}'] = r.get('obs','')
                fila += 1
            ws['K37'] = sum(float(r.get('m2_rollo', 0) or 0) for r in registros)
            ws['Q37'] = sum(float(r.get('st_m2', 0) or 0) for r in registros)
            ws['U37'] = sum(float(r.get('ct_m2', 0) or 0) for r in registros)

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'{proto}.xlsx'
        )

    except Exception as e:
        return jsonify({'error': str(e)}), 500


if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port)
